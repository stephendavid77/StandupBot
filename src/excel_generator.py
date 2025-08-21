import openpyxl
from openpyxl.styles import Font, Alignment

class ExcelGenerator:
    def __init__(self):
        pass

    def generate_excel(self, results, output_excel_path):
        wb = openpyxl.Workbook()
        
        # Remove the default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        self._create_summary_sheet(wb, results)
        self._create_hygiene_sheet(wb, results)
        self._create_workload_sheet(wb, results)

        wb.save(output_excel_path)

    def _create_summary_sheet(self, wb, results):
        ws = wb.create_sheet("Sprint Summary")
        sprint_info = results["sprint_info"]
        metrics = results["metrics"]

        ws.append(["Sprint Name", sprint_info["name"]])
        ws.append(["Sprint Goal", sprint_info["goal"]])
        ws.append(["Start Date", sprint_info["start_date"]])
        ws.append(["End Date", sprint_info["end_date"]])
        ws.append([])
        ws.append(["Total Issues", metrics["total_issues"]])
        ws.append(["Issues Done", metrics["status_counts"]["Done"]])
        ws.append(["Progress Pct", f"{metrics['progress_pct']:.2f}%"])
        ws.append(["Total Story Points", metrics["total_story_points"]])
        ws.append(["Story Points Done", metrics["story_points_done"]])

        self._format_sheet(ws)

    def _create_hygiene_sheet(self, wb, results):
        ws = wb.create_sheet("Sprint Hygiene")
        hygiene = results["hygiene"]

        ws.append(["Unassigned Issues", hygiene["unassigned_issues"]["count"]])
        ws.append(["Unestimated Issues", hygiene["unestimated_issues"]["count"]])
        ws.append(["Issues without Fix Versions", hygiene["issues_without_fix_versions"]["count"]])
        ws.append([])
        ws.append(["Unassigned Issues Keys"]) 
        ws.append(hygiene["unassigned_issues"]["keys"])
        ws.append(["Unestimated Issues Keys"])
        ws.append(hygiene["unestimated_issues"]["keys"])
        ws.append(["Issues without Fix Versions Keys"])
        ws.append(hygiene["issues_without_fix_versions"]["keys"])

        self._format_sheet(ws)

    def _create_workload_sheet(self, wb, results):
        ws = wb.create_sheet("Workload")
        workload = results["workload"]

        ws.append(["Assignee", "Total Points"])
        for assignee, points in workload.items():
            ws.append([assignee, points["total_points"]])

        self._format_sheet(ws)

    def _format_sheet(self, ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Bold the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
